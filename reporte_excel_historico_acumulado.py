# -*- coding: utf-8 -*-
"""
MOTOR DE CONSOLIDACIÓN HISTÓRICA EN EXCEL - DITIC
Archivo: reporte_excel_historico_acumulado.py
Versión: V82.9.34 (Refactorización a Código Limpio en Español)
Autor: Mg. Fredy Alejandro Sarmiento Torres
Lógica: Combina datos históricos (Excel existente) con datos nuevos (memoria) y genera un libro multicapa.
"""

# Importamos librerías del sistema y recolección de basura
import os
import time
import datetime
import logging
import traceback
import gc

# Importamos pandas para la manipulación de dataframes masivos
import pandas as pd
# Importamos Path para el manejo de rutas
from pathlib import Path

# Importamos utilidades de openpyxl para crear y darle diseño profesional a los libros de Excel
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule

# Obtenemos la instancia del logger configurada previamente
registrador_eventos = logging.getLogger()

class LectorHistoricoInteligente:
    """Clase encargada de leer el archivo Excel anterior de manera resiliente (Soporta cambios de formato antiguos)."""
    
    @staticmethod
    def leer_historico_seguro(ruta_archivo_excel):
        """Intenta leer la hoja de 'Datos Procesados' detectando en qué fila inician los encabezados reales."""
        ruta_fisica = Path(ruta_archivo_excel)
        # Si el archivo no existe (ej. es el primer mes que se procesa el área), retornamos un DataFrame vacío
        if not ruta_fisica.exists(): 
            return pd.DataFrame()
        
        try:
            # Intento 1: Leemos las primeras 20 filas buscando la fila exacta donde están los encabezados reales
            dataframe_temporal = pd.read_excel(ruta_fisica, sheet_name='Datos Procesados', header=None, nrows=20)
            
            # Iteramos sobre esas filas
            for indice_fila, fila_datos in dataframe_temporal.iterrows():
                # Limpiamos espacios en blanco de todos los valores de la fila
                valores_fila_limpios = [str(valor).strip() for valor in fila_datos.values]
                # Si encontramos estas dos columnas clave, sabemos que ESA es la fila de encabezados
                if 'Año' in valores_fila_limpios and 'Métrica' in valores_fila_limpios:
                    # Volvemos a leer todo el archivo, pero diciéndole a pandas que arranque en ese índice exacto
                    return pd.read_excel(ruta_fisica, sheet_name='Datos Procesados', header=indice_fila)
        except Exception: 
            # Si la hoja no existe o el archivo está corrupto, lo ignoramos silenciosamente
            pass
        
        # En caso de fallo total de lectura, retornamos DataFrame vacío
        return pd.DataFrame()

class GeneradorReporteAcumulado:
    """Reportero de Alto Nivel: Fusiona los datos, genera múltiples vistas (Raw, Procesado, YoY) y aplica formato Institucional."""

    # Diccionario para renombrar las columnas crudas de SQL a nombres más cortos y amigables en la hoja de Auditoría
    MAPA_COLUMNAS_AUDITORIA = {
        'encuestadoId': 'ID_Enc', 'Año': 'Año', 'Mes': 'Mes', 'respuestaFch': 'Fecha',
        'preguntaDescripcion': 'Pregunta', 'encuestaNombre': 'Encuesta', 'servicioNombre': 'Servicio',
        'sedeNombre': 'Sede', 'areaNombre': 'Area', 'encuestadoApellidos': 'Apellidos',
        'encuestadoNombres': 'Nombres', 'encuestadoCelular': 'Celular', 'encuestadoEmail': 'Email',
        'tipoPreguntaId': 'T_Preg', 'consecutivo': 'Consec', 'preguntaId': 'ID_P',
        'encuestaId': 'ID_E', 'idAreaServicio': 'ID_AS', 'idServicioEncuesta': 'ID_SE',
        'idSedeEncuesta': 'ID_SedeE', 'respuestaId': 'ID_Resp', 'Atención': 'Att',
        'Comunicación y acceso': 'Com_Acc', 'Eficiencia': 'Eficiencia', 'NPS': 'NPS',
        'Resolución de la necesidad': 'Resol', 'Tiempo de respuesta': 'T_Resp',
        'preguntaSinIndicador': 'P_Sin_Ind', 'NPS_Numerico': 'NPS_N', 'Atencion_Numerica': 'Att_N'
    }

    @classmethod
    def procesar_historico_mensual(cls, dataframe_procesado_nuevo, dataframe_crudo_sql, nombre_area, anio_proceso, mes_proceso, ruta_guardado_local):
        """Método principal que orquesta la lectura, fusión, diseño y guardado del archivo de Excel Acumulado."""
        try:
            # 1. Recuperación del histórico existente
            dataframe_historico = LectorHistoricoInteligente.leer_historico_seguro(ruta_guardado_local)
            
            if not dataframe_historico.empty:
                # REGLA DE ORO: Borramos los datos del mes/año que estamos procesando en el histórico.
                # Esto evita duplicidad si corremos el script dos veces para el mismo mes.
                dataframe_historico = dataframe_historico[~((dataframe_historico['Año'] == anio_proceso) & (dataframe_historico['Mes'] == mes_proceso))]
                
                # Fusionamos (Append) el histórico limpio con los datos recién procesados de SQL
                dataframe_final_fusionado = pd.concat([dataframe_historico, dataframe_procesado_nuevo], ignore_index=True)
                registrador_eventos.info(f"📈 Fusión Exitosa: {len(dataframe_historico)} registros previos + {len(dataframe_procesado_nuevo)} registros nuevos.")
            else:
                # Si no había histórico, el archivo final solo contendrá los datos del mes actual
                dataframe_final_fusionado = dataframe_procesado_nuevo

            # 2. Creación del Libro Maestro de Excel
            libro_excel = Workbook()
            # Eliminamos la hoja 'Sheet1' que viene por defecto
            libro_excel.remove(libro_excel.active)

            # --- HOJA 0: SQL RAW DATA (Transparencia Total - Auditoría) ---
            hoja_auditoria = libro_excel.create_sheet("SQL Raw Data", 0)
            cls._construir_hoja_auditoria(hoja_auditoria, dataframe_crudo_sql, nombre_area)

            # --- HOJA 1: DATOS PROCESADOS (Sábana de Métricas limpias) ---
            hoja_procesados = libro_excel.create_sheet("Datos Procesados")
            cls._construir_hoja_procesados(hoja_procesados, dataframe_final_fusionado, nombre_area)

            # --- HOJA 2: TENDENCIA TEMPORAL (Evolución Mes a Mes) ---
            hoja_tendencias = libro_excel.create_sheet("Tendencia Temporal")
            cls._construir_hoja_tendencia(hoja_tendencias, dataframe_final_fusionado)

            # --- HOJA 3: COMPARATIVO ANUAL (Análisis Year-over-Year) ---
            hoja_comparativa = libro_excel.create_sheet("Comparativo Anual")
            cls._construir_hoja_comparativa_anual(hoja_comparativa, dataframe_final_fusionado)

            # 3. Guardado en Disco Físico
            libro_excel.save(ruta_guardado_local)
            
            # Forzamos al recolector de basura a limpiar la memoria RAM usada por los dataframes pesados
            gc.collect()
            return True

        except Exception as excepcion_capturada:
            registrador_eventos.error(f"❌ Error en Motor de Acumulado: {excepcion_capturada}\n{traceback.format_exc()}")
            return False

    @classmethod
    def _construir_hoja_auditoria(cls, hoja_trabajo, dataframe_crudo, nombre_area):
        """Escribe los datos crudos de la vista SQL en la hoja de Excel."""
        # Colocamos los logos y cabeceras
        cls._estampar_encabezado_institucional(hoja_trabajo, f"AUDITORÍA DE DATOS CRUDOS SQL: {nombre_area}")
        
        # Renombramos las columnas usando nuestro mapa de diccionario
        dataframe_renombrado = dataframe_crudo.copy().rename(columns=cls.MAPA_COLUMNAS_AUDITORIA)
        
        # Iteramos sobre el dataframe y escribimos fila por fila en Excel
        for fila_datos in dataframe_to_rows(dataframe_renombrado, index=False, header=True):
            hoja_trabajo.append(fila_datos)
        
        # Aplicamos el estilo de tabla (colores alternos y filtros) a partir de la fila 5
        cls._aplicar_estilo_tabla_institucional(hoja_trabajo, "TablaAuditoriaSQL", fila_inicio=5)

    @classmethod
    def _construir_hoja_procesados(cls, hoja_trabajo, dataframe_procesado, nombre_area):
        """Escribe la sábana procesada (melted) y aplica semaforización."""
        cls._estampar_encabezado_institucional(hoja_trabajo, f"HISTÓRICO DE DATOS PROCESADOS: {nombre_area}")
        
        for fila_datos in dataframe_to_rows(dataframe_procesado, index=False, header=True):
            hoja_trabajo.append(fila_datos)
            
        # Activamos semaforo=True para que pinte de verde/amarillo/rojo la columna de indicadores
        cls._aplicar_estilo_tabla_institucional(hoja_trabajo, "TablaHistorica", fila_inicio=5, activar_semaforo=True)

    @classmethod
    def _construir_hoja_tendencia(cls, hoja_trabajo, dataframe_procesado):
        """Agrupa los datos por Año y Mes para mostrar la evolución del promedio general."""
        cls._estampar_encabezado_institucional(hoja_trabajo, "TENDENCIA MENSUAL DE PERCEPCIÓN DE SERVICIOS")
        
        # Agrupamos calculando cuántas encuestas hubo y el promedio del indicador
        tendencia_agrupada = dataframe_procesado.groupby(['Año', 'Mes'])['Indicador_0_100'].agg(['count', 'mean']).reset_index()
        tendencia_agrupada.columns = ['Año', 'Mes', 'Total Encuestas', 'Promedio Percepción']
        
        # Dividimos entre 100 para que Excel pueda aplicar el formato de porcentaje (0.80 -> 80%)
        tendencia_agrupada['Promedio Percepción'] = tendencia_agrupada['Promedio Percepción'] / 100
        
        for fila_datos in dataframe_to_rows(tendencia_agrupada, index=False, header=True):
            hoja_trabajo.append(fila_datos)
        
        # Aplicar formato de Porcentaje (0.0%) a la columna 4 (Promedio Percepción)
        for num_fila in range(6, hoja_trabajo.max_row + 1):
            hoja_trabajo.cell(num_fila, 4).number_format = '0.0%'
        
        cls._aplicar_estilo_tabla_institucional(hoja_trabajo, "TablaTendencia", fila_inicio=5)

    @classmethod
    def _construir_hoja_comparativa_anual(cls, hoja_trabajo, dataframe_procesado):
        """Genera una matriz dinámica (Pivot Table) para comparar métricas año contra año."""
        cls._estampar_encabezado_institucional(hoja_trabajo, "MATRIZ COMPARATIVA ANUAL POR MÉTRICA")
        try:
            # Construimos la Pivot: Filas (Métrica), Columnas (Años), Valores (Promedio)
            tabla_pivote = dataframe_procesado.pivot_table(index='Métrica', columns='Año', values='Indicador_0_100', aggfunc='mean') / 100
            
            # Extraemos y escribimos los encabezados de la Pivot Table
            encabezados_pivot = ['Métrica'] + [str(columna) for columna in tabla_pivote.columns]
            hoja_trabajo.append(encabezados_pivot)
            
            # Escribimos los datos de la Pivot
            for nombre_metrica, fila_datos in tabla_pivote.iterrows():
                hoja_trabajo.append([nombre_metrica] + list(fila_datos.values))
            
            # Aplicamos formato de porcentaje con 1 decimal (0.0%) a todas las celdas numéricas de la matriz
            for num_fila in range(6, hoja_trabajo.max_row + 1):
                for num_columna in range(2, hoja_trabajo.max_column + 1):
                    hoja_trabajo.cell(num_fila, num_columna).number_format = '0.1%'
                    
            cls._aplicar_estilo_tabla_institucional(hoja_trabajo, "TablaComparativaAnual", fila_inicio=5)
        except Exception:
            # Si hay muy pocos datos o falla la pivotación, dejamos un mensaje claro en la hoja
            hoja_trabajo.append(["Nota: Datos insuficientes para generar matriz comparativa anual."])

    @staticmethod
    def _estampar_encabezado_institucional(hoja_trabajo, titulo_hoja):
        """Inserta el diseño estándar DITIC en las primeras 4 filas de cualquier hoja de Excel."""
        hoja_trabajo['A1'] = "UNIVERSIDAD DEL ROSARIO - DITIC"
        hoja_trabajo['A1'].font = Font(bold=True, size=14, color="AF2024") # Color Institucional Rojo
        hoja_trabajo['A2'] = titulo_hoja
        hoja_trabajo['A2'].font = Font(bold=True, size=12, color="1F4E78") # Color Institucional Azul
        hoja_trabajo['A3'] = f"Reporte Generado: {datetime.datetime.now():%Y-%m-%d %H:%M} | Responsable: Fredy A. Sarmiento"
        hoja_trabajo['A3'].font = Font(italic=True, size=9)
        hoja_trabajo.append([]) # Dejamos la fila 4 vacía como espacio separador antes de la tabla

    @staticmethod
    def _aplicar_estilo_tabla_institucional(hoja_trabajo, prefijo_nombre_tabla, fila_inicio, activar_semaforo=False):
        """Convierte un rango de datos en una 'Tabla Inteligente' de Excel (con filtros y estilos predefinidos)."""
        max_columna, max_fila = hoja_trabajo.max_column, hoja_trabajo.max_row
        
        # Si no hay datos debajo de la fila de inicio, cancelamos
        if max_fila <= fila_inicio: 
            return
        
        # Definimos el rango exacto de la tabla (ej. A5:J100)
        rango_referencia = f"A{fila_inicio}:{get_column_letter(max_columna)}{max_fila}"
        
        # Creamos el objeto Tabla, agregando un número aleatorio al nombre para evitar conflictos en Excel
        tabla_excel = Table(displayName=f"{prefijo_nombre_tabla}_{int(time.time()*100)%100000}", ref=rango_referencia)
        # Aplicamos un estilo visual azulado de bandas medias estándar en Excel
        tabla_excel.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        hoja_trabajo.add_table(tabla_excel)
        
        # Sobrescribimos el diseño de la cabecera (Fila 5) con los colores exactos de la DITIC
        relleno_cabecera = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        fuente_cabecera = Font(bold=True, color="FFFFFF")
        
        for indice_col in range(1, max_columna + 1):
            celda = hoja_trabajo.cell(fila_inicio, indice_col)
            celda.fill = relleno_cabecera
            celda.font = fuente_cabecera
            celda.alignment = Alignment(horizontal='center')

        # Auto-ajustamos el ancho de todas las columnas a 18 píxeles para mejor lectura
        for columna_tupla in hoja_trabajo.columns:
            letra_columna = get_column_letter(columna_tupla[0].column)
            hoja_trabajo.column_dimensions[letra_columna].width = 18

        # Si se solicita, aplicamos una regla de Formato Condicional (Semáforo de colores 0 a 100) a la última columna
        if activar_semaforo:
            letra_ultima_columna = get_column_letter(max_columna)
            # Aplicamos regla ColorScaleRule (Rojo=0, Amarillo=75, Verde=100)
            hoja_trabajo.conditional_formatting.add(
                f"{letra_ultima_columna}{fila_inicio+1}:{letra_ultima_columna}{max_fila}", 
                ColorScaleRule(start_type='num', start_value=0, start_color='F8696B',
                               mid_type='num', mid_value=75, mid_color='FFEB84',
                               end_type='num', end_value=100, end_color='63BE7B')
            )