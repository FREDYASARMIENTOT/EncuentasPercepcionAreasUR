
# -*- coding: utf-8 -*-
"""
MOTOR DE CONSOLIDACIÓN HISTÓRICA - DITIC
Versión: V82.9.9 Titanium Elite (Executive Design & Raw Auditor)
Autor: Mg. Fredy Alejandro Sarmiento Torres
"""

import os
import time
import datetime
import logging
import traceback
import gc
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule

from config import Config

logger = logging.getLogger()

class HistoricalDataReader:
    """Lector Resiliente con Scan Dinámico (Legacy Support V77)."""
    @staticmethod
    def read_smart(filepath):
        path = Path(filepath)
        if not path.exists(): return pd.DataFrame()
        
        # Estrategias de lectura para evitar rupturas por cambios de formato
        try:
            # Intento 1: Hoja estándar de datos procesados
            df = pd.read_excel(path, sheet_name='Datos Procesados', header=None, nrows=20)
            for idx, row in df.iterrows():
                row_vals = [str(v).strip() for v in row.values]
                if 'Año' in row_vals and 'Métrica' in row_vals:
                    return pd.read_excel(path, sheet_name='Datos Procesados', header=idx)
        except: pass
        return pd.DataFrame()

class ExcelAcumuladoReporter:
    """Reportero de Alto Nivel: Fusiona V63.0 (Estética) con V82.9 (Auditoría)."""

    # Diccionario de nombres mínimos para la hoja RAW (Auditoría)
    RAW_MAP = {
        'encuestadoId': 'ID_Enc', 'Año': 'Año', 'Mes': 'Mes', 'respuestaFch': 'Fecha',
        'preguntaDescripcion': 'Pregunta', 'encuestaNombre': 'Encuesta', 'servicioNombre': 'Servicio',
        'sedeNombre': 'Sede', 'areaNombre': 'Area', 'encuestadoApellidos': 'Apellidos',
        'encuestadoNombres': 'Nombres', 'encuestadoCelular': 'Celular', 'encuestadoEmail': 'Email',
        'tipoPreguntaId': 'T_Preg', 'consecutivo': 'Consec', 'preguntaId': 'ID_P',
        'encuestaId': 'ID_E', 'idAreaServicio': 'ID_AS', 'idServicioEncuesta': 'ID_SE',
        'idSedeEncuesta': 'ID_SedeE', 'respuestaId': 'ID_Resp', 'Atención': 'Att',
        'Comunicación y acceso': 'Com_Acc', 'Eficiencia': 'Eficiencia', 'NPS': 'NPS',
        'Resolución de la necesidad': 'Resol', 'Tiempo de respuesta': 'T_Resp',
        'preguntaSinIndicador': 'P_Sin_Ind', 'NPS_Numerico': 'NPS_N', 'Atencion_Numerica': 'Att_N'
    }

    @classmethod
    def procesar(cls, df_proc, df_raw, area, anio, mes, ruta):
        """Genera el acumulado con el ADN de la V63.0 y la transparencia de la V82.9."""
        try:
            # 1. Recuperación del histórico (Smart Read)
            df_old = HistoricalDataReader.read_smart(ruta)
            
            if not df_old.empty:
                # Limpieza de duplicados por mes/año para re-procesamiento
                df_old = df_old[~((df_old['Año'] == anio) & (df_old['Mes'] == mes))]
                df_final_proc = pd.concat([df_old, df_proc], ignore_index=True)
                logger.info(f"📈 Fusión: {len(df_old)} previos + {len(df_proc)} nuevos.")
            else:
                df_final_proc = df_proc

            # 2. Creación del Libro Maestro
            wb = Workbook()
            wb.remove(wb.active)

            # --- HOJA 0: SQL RAW DATA (Transparencia Total) ---
            ws_raw = wb.create_sheet("SQL Raw Data", 0)
            cls._build_raw_sheet(ws_raw, df_raw, area)

            # --- HOJA 1: DATOS PROCESADOS (Sábana de Métricas) ---
            ws_proc = wb.create_sheet("Datos Procesados")
            cls._build_proc_sheet(ws_proc, df_final_proc, area)

            # --- HOJA 2: TENDENCIA TEMPORAL (Evolución) ---
            ws_trend = wb.create_sheet("Tendencia Temporal")
            cls._build_trend_sheet(ws_trend, df_final_proc)

            # --- HOJA 3: COMPARATIVO ANUAL (YoY Analysis) ---
            ws_yoy = wb.create_sheet("Comparativo Anual")
            cls._build_yoy_sheet(ws_yoy, df_final_proc)

            # 3. Persistencia
            wb.save(ruta)
            gc.collect()
            return True

        except Exception as e:
            logger.error(f"❌ Error en Motor Acumulado V82.9.9: {e}\n{traceback.format_exc()}")
            return False

    @classmethod
    def _build_raw_sheet(cls, ws, df, area):
        """Implementa el volcado total de la vista SQL (Raw View)."""
        cls._estampar_encabezado(ws, f"AUDITORÍA RAW SQL: {area}")
        
        # Columnas optimizadas
        df_temp = df.copy().rename(columns=cls.RAW_MAP)
        
        for r in dataframe_to_rows(df_temp, index=False, header=True):
            ws.append(r)
        
        cls._aplicar_estilo_titan(ws, "TblRawSQL", row=5)

    @classmethod
    def _build_proc_sheet(cls, ws, df, area):
        """Sábana procesada con métricas melteadas."""
        cls._estampar_encabezado(ws, f"HISTÓRICO PROCESADO: {area}")
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        cls._aplicar_estilo_titan(ws, "TblHistorica", row=5, semaforo=True)

    @classmethod
    def _build_trend_sheet(cls, ws, df):
        """Evolución de promedios por mes y año."""
        cls._estampar_encabezado(ws, "TENDENCIA MENSUAL DE PERCEPCIÓN")
        trend = df.groupby(['Año', 'Mes'])['Indicador_0_100'].agg(['count', 'mean']).reset_index()
        trend.columns = ['Año', 'Mes', 'Encuestas', 'Promedio Percepción']
        trend['Promedio Percepción'] = trend['Promedio Percepción'] / 100
        
        for r in dataframe_to_rows(trend, index=False, header=True):
            ws.append(r)
        
        # Formato %
        for r in range(6, ws.max_row + 1):
            ws.cell(r, 4).number_format = '0.0%'
        
        cls._aplicar_estilo_titan(ws, "TblTendencia", row=5)

    @classmethod
    def _build_yoy_sheet(cls, ws, df):
        """Matriz Comparativa Inter-anual (V63.0 Elite)."""
        cls._estampar_encabezado(ws, "COMPARATIVO ANUAL POR MÉTRICA")
        try:
            pivot = df.pivot_table(index='Métrica', columns='Año', values='Indicador_0_100', aggfunc='mean') / 100
            headers = ['Métrica'] + [str(c) for c in pivot.columns]
            ws.append(headers)
            for idx, row in pivot.iterrows():
                ws.append([idx] + list(row.values))
            
            # Formato %
            for r in range(6, ws.max_row + 1):
                for c in range(2, ws.max_column + 1):
                    ws.cell(r, c).number_format = '0.1%'
            cls._aplicar_estilo_titan(ws, "TblYoY", row=5)
        except:
            ws.append(["Nota: Datos insuficientes para generar matriz comparativa."])

    @staticmethod
    def _estampar_encabezado(ws, titulo):
        """Diseño Institucional V63.0."""
        ws['A1'] = "UNIVERSIDAD DEL ROSARIO - DITIC"
        ws['A1'].font = Font(bold=True, size=14, color="AF2024")
        ws['A2'] = titulo
        ws['A2'].font = Font(bold=True, size=12, color="1F4E78")
        ws['A3'] = f"Generado: {datetime.datetime.now():%Y-%m-%d %H:%M} | Responsable: Fredy A. Sarmiento"
        ws['A3'].font = Font(italic=True, size=9)
        ws.append([]) # Fila 4 vacía

    @staticmethod
    def _aplicar_estilo_titan(ws, nombre, row, semaforo=False):
        """Aplica formato de tabla profesional y autoajuste."""
        mc, mr = ws.max_column, ws.max_row
        if mr <= row: return
        
        ref = f"A{row}:{get_column_letter(mc)}{mr}"
        tab = Table(displayName=f"{nombre}_{int(time.time()*100)%100000}", ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(tab)
        
        # Formato de cabecera manual (V63.0 Style)
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for i in range(1, mc + 1):
            cell = ws.cell(row, i)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Auto-ajuste de columnas
        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = 18

        if semaforo:
            cl = get_column_letter(mc)
            ws.conditional_formatting.add(f"{cl}{row+1}:{cl}{mr}", 
                ColorScaleRule(start_type='num', start_value=0, start_color='F8696B',
                               mid_type='num', mid_value=75, mid_color='FFEB84',
                               end_type='num', end_value=100, end_color='63BE7B'))