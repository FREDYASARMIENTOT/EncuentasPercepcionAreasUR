# -*- coding: utf-8 -*-
"""
MOTOR DE TRANSFORMACIÓN DE DATOS Y GENERACIÓN DE DASHBOARDS MENSUALES - DITIC
Archivo: reporte_excel_ejecutivo_mensual.py
Versión: V82.9.34 (Refactorización a Código Limpio en Español)
Autor: Mg. Fredy Alejandro Sarmiento Torres
Lógica: Procesa los datos crudos de SQL (Melt), analiza sentimientos y genera el libro Excel con KPIs visuales.
"""

# Importamos librerías del sistema y control de memoria
import os
import time
import datetime
import gc
import traceback

# Importamos pandas para transformación de datos y numpy para manejo de valores nulos
import pandas as pd
import numpy as np

# Importamos utilidades de openpyxl para construir y dar diseño al archivo Excel
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image as XLImage

# Importamos Pillow (PIL) para manipular imágenes en memoria antes de pasarlas a Excel
from PIL import Image as PILImage
from io import BytesIO

# Configuración defensiva de Matplotlib para operar en servidores de Windows (sin monitor)
try:
    import matplotlib
    # Forzamos el backend 'Agg' que no intenta abrir ventanas emergentes y guarda directo a archivo/memoria
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.colors as mcolors
    DISPONIBILIDAD_MATPLOTLIB = True
except ImportError:
    # Si la librería no está instalada, el script no colapsa, solo desactiva los gráficos
    DISPONIBILIDAD_MATPLOTLIB = False

# Importamos nuestra configuración centralizada
from configuracion_sistema_encuestas import ConfiguracionSistema

# ==========================================================================
# 🧠 CLASE 1: MOTOR DE TRANSFORMACIÓN DE DATOS
# ==========================================================================
class TransformadorDatos:
    """Clase encargada de convertir la tabla plana de SQL (cruda) en un formato analítico tabular estructurado."""

    @staticmethod
    def calcular_indicador_desempeno(valor_respuesta):
        """
        Motor Analítico de Sentimientos (Sentiment Engine).
        Estandariza cualquier tipo de respuesta a una escala porcentual de 0 a 100.
        """
        # Si la respuesta está vacía, se asume neutralidad (75.0) para no castigar el indicador injustamente
        if pd.isna(valor_respuesta): 
            return 75.0 
        
        # Limpiamos espacios y pasamos a mayúsculas para facilitar la comparación de texto
        texto_limpio = str(valor_respuesta).strip().upper()
        
        try:
            # Intentamos convertir la respuesta a número, reemplazando comas por puntos decimales
            numero_extraido = float(texto_limpio.replace(',', '.'))
            
            # Lógica 1: Escalas de 1 a 5 (ej. Estrellas). 5 = 100%, 1 = 0%
            if 1 <= numero_extraido <= 5: 
                return (numero_extraido - 1) * 25.0
            
            # Lógica 2: Escalas de 1 a 10 (ej. NPS clásico). 10 = 100%, 1 = 0%
            if 5 < numero_extraido <= 10: 
                return (numero_extraido - 1) * (100.0 / 9.0)
            
            # Si es un número en otra escala (ej. ya viene sobre 100), lo dejamos pasar
            return numero_extraido 
        except Exception: 
            # Si no se pudo convertir a número, asumimos que es una respuesta de texto libre (Open Text)
            pass

        # Traemos el diccionario maestro de palabras clave positivas/negativas desde la configuración
        diccionario_palabras = ConfiguracionSistema.PALABRAS_CLAVE_SENTIMIENTO
        
        # Si el texto contiene alguna palabra negativa, asignamos 0%
        if any(palabra in texto_limpio for palabra in diccionario_palabras['NEGATIVO']): 
            return 0.0
        # Si contiene alguna palabra positiva, asignamos 100%
        if any(palabra in texto_limpio for palabra in diccionario_palabras['POSITIVO']): 
            return 100.0
        
        # Si no detecta ni positivo ni negativo, asume neutralidad y otorga 75%
        return 75.0

    @classmethod
    def procesar_datos_crudos(cls, dataframe_crudo):
        """Transforma las columnas anchas de métricas en filas profundas (Operación Melt/Unpivot)."""
        # Lista exacta de las columnas de SQL que representan métricas a evaluar
        columnas_metricas = ['Atención', 'Comunicación y acceso', 'Eficiencia', 'NPS', 
                             'Resolución de la necesidad', 'Tiempo de respuesta']
        
        # Filtramos solo las métricas que realmente vienen en la consulta SQL
        metricas_existentes = [columna for columna in columnas_metricas if columna in dataframe_crudo.columns]
        # Todo lo demás (ID, Nombre, Fecha, etc.) se considera columnas fijas o de contexto
        columnas_contexto = [columna for columna in dataframe_crudo.columns if columna not in metricas_existentes]
        
        # Unpivot: Convertimos las columnas de métricas en dos nuevas columnas ('Métrica' y 'Valor')
        dataframe_transformado = dataframe_crudo.melt(
            id_vars=columnas_contexto, 
            value_vars=metricas_existentes, 
            var_name='Métrica', 
            value_name='Valor'
        ).dropna(subset=['Valor']) # Eliminamos filas donde la persona no evaluó esa métrica en particular

        # Aplicamos nuestro motor de sentimientos a la columna 'Valor' para generar el indicador numérico
        dataframe_transformado['Indicador_0_100'] = dataframe_transformado['Valor'].apply(cls.calcular_indicador_desempeno).astype(float)
        
        # Mapeo de limpieza: Capitalizamos los nombres clave para que Power BI y los reportes se vean elegantes
        mapa_nombres_limpios = {'areanombre': 'Área', 'sedenombre': 'Sede', 'servicionombre': 'Servicio'}
        dataframe_transformado.columns = [mapa_nombres_limpios.get(columna.lower(), columna) for columna in dataframe_transformado.columns]
        
        return dataframe_transformado

# ==========================================================================
# 📊 CLASE 2: GENERADOR DE REPORTES EXCEL
# ==========================================================================
class GeneradorReporteEjecutivoMensual:
    """Clase encargada de construir el libro Excel mensual, distribuyendo la información en pestañas jerárquicas."""

    # Mapa de compresión de nombres para la hoja de auditoría, ahorrando espacio horizontal
    MAPA_COLUMNAS_AUDITORIA = {
        'encuestadoId': 'ID_Enc', 'Año': 'Año', 'Mes': 'Mes', 'respuestaFch': 'Fecha',
        'preguntaDescripcion': 'Pregunta', 'servicioNombre': 'Servicio',
        'sedeNombre': 'Sede', 'areaNombre': 'Area', 'consecutivo': 'Consec', 
        'respuestaId': 'ID_Resp'
    }

    @classmethod
    def construir_libro_excel(cls, dataframe_procesado, dataframe_crudo_sql, nombre_area, anio, mes, ruta_guardado):
        """Orquestador maestro del diseño del Excel. Crea las hojas en el orden requerido por la auditoría."""
        libro_excel = Workbook()
        libro_excel.remove(libro_excel.active)
        
        # Validamos que tengamos datos procesados, de lo contrario creamos un DataFrame vacío con la columna clave
        dataframe_seguro = dataframe_procesado.copy() if not dataframe_procesado.empty else pd.DataFrame(columns=['Indicador_0_100'])

        # --- ORDEN DE PESTAÑAS EXIGIDO INSTITUCIONALMENTE ---
        # 1. POSICIÓN 0: SQL RAW (Obligatorio para auditorías de datos crudos)
        cls._construir_pestana_datos_crudos(libro_excel, dataframe_crudo_sql, nombre_area)
        
        # 2. POSICIÓN 1: DASHBOARD EJECUTIVO (Resumen visual rápido)
        cls._construir_pestana_dashboard_visual(libro_excel, dataframe_seguro, mes, anio)

        # 3. PESTAÑAS TÉCNICAS (Detalle profundo)
        cls._construir_pestana_sabana_procesada(libro_excel, dataframe_seguro, nombre_area)
        cls._construir_pestana_resumen_volumetria(libro_excel, dataframe_seguro)
        cls._construir_pestana_calidad_analitica(libro_excel, dataframe_seguro)
        cls._construir_pestana_comentarios_abiertos(libro_excel, dataframe_seguro)

        try:
            # Guardamos el archivo físico
            libro_excel.save(ruta_guardado)
            gc.collect() # Liberamos memoria RAM
            # Retornamos éxito y enviamos el conteo de encuestas únicas (consecutivo) para que salga en el correo
            return True, {"cons": dataframe_seguro['consecutivo'].nunique() if 'consecutivo' in dataframe_seguro.columns else 0}
        except Exception as excepcion_capturada:
            print(f"❌ Error Crítico guardando el reporte mensual Excel: {excepcion_capturada}")
            return False, {}

    # ----------------------------------------------------------------------
    # 📑 CONSTRUCCIÓN INTERNA DE PESTAÑAS (MÉTODOS OCULTOS)
    # ----------------------------------------------------------------------

    @classmethod
    def _construir_pestana_datos_crudos(cls, libro_excel, dataframe, nombre_area):
        """Crea la hoja índice 0 con la descarga literal de la base de datos SQL."""
        # Forzamos que sea la primera pestaña visible (índice 0)
        hoja_trabajo = libro_excel.create_sheet("SQL Raw Data", 0) 
        cls._aplicar_cabecera_institucional(hoja_trabajo, f"AUDITORÍA FUENTE DWH: {nombre_area}")
        
        # Renombramos y escribimos de forma segura
        dataframe_renombrado = dataframe.copy().rename(columns=cls.MAPA_COLUMNAS_AUDITORIA)
        cls._escribir_dataframe_seguro(hoja_trabajo, dataframe_renombrado)
        
        # Aplicamos la tabla con estilo desde la fila 5
        rango_tabla = f"A5:{get_column_letter(hoja_trabajo.max_column)}{hoja_trabajo.max_row}"
        cls._crear_tabla_dinamica_excel(hoja_trabajo, "TblAuditoriaRaw", rango_tabla)

    @classmethod
    def _construir_pestana_dashboard_visual(cls, libro_excel, dataframe, mes, anio):
        """Prepara el lienzo para el Dashboard de KPIs y gráficos."""
        hoja_trabajo = libro_excel.create_sheet("Dashboard Visual")
        cls._aplicar_cabecera_institucional(hoja_trabajo, f"TABLERO EJECUTIVO {mes}/{anio}")
        
        # KPI Principal: Puntuación Global Promedio
        hoja_trabajo['A4'] = "PUNTUACIÓN GLOBAL:"
        # Dividimos entre 100 para aplicar formato de porcentaje nativo en Excel
        hoja_trabajo['B4'] = dataframe['Indicador_0_100'].mean() / 100
        hoja_trabajo['B4'].number_format = '0.0%'
        hoja_trabajo['B4'].font = Font(bold=True, size=14, color="1F4E78")
        
        # Espacio reservado para la inyección de gráficos de Matplotlib (si aplican)
        celda_mensaje = hoja_trabajo.cell(6, 1, "💡 Visualizaciones gráficas y analítica listas para su revisión.")
        celda_mensaje.font = Font(italic=True)

    @classmethod
    def _construir_pestana_sabana_procesada(cls, libro_excel, dataframe, nombre_area):
        """Escribe la tabla limpia con el resultado numérico 0-100 para cada fila."""
        hoja_trabajo = libro_excel.create_sheet("Datos Procesados")
        cls._aplicar_cabecera_institucional(hoja_trabajo, f"SÁBANA DE MÉTRICAS ANALÍTICAS: {nombre_area}")
        
        # Seleccionamos solo las columnas relevantes
        columnas_deseadas = [col for col in ['Sede', 'Servicio', 'Métrica', 'Valor', 'Indicador_0_100'] if col in dataframe.columns]
        cls._escribir_dataframe_seguro(hoja_trabajo, dataframe[columnas_deseadas])
        
        # Rango hasta la columna E y activamos el semáforo en esa columna
        cls._crear_tabla_dinamica_excel(hoja_trabajo, "TblDatosProcesados", f"A5:E{hoja_trabajo.max_row}", columna_semaforo="E")

    @classmethod
    def _construir_pestana_resumen_volumetria(cls, libro_excel, dataframe):
        """Agrupa y cuenta cuántas encuestas se recibieron por sede y servicio."""
        hoja_trabajo = libro_excel.create_sheet("Totalización Volumétrica")
        cls._aplicar_cabecera_institucional(hoja_trabajo, "RESUMEN DE VOLUMETRÍA E IMPACTO")
        if dataframe.empty: return
        
        # Agrupamos calculando conteo total y promedio general
        agrupacion_volumen = dataframe.groupby(['Sede', 'Servicio']).agg(
            Cantidad_Evaluaciones=('Métrica','count'), 
            Promedio_Indicador=('Indicador_0_100','mean')
        ).reset_index()
        
        cls._escribir_dataframe_seguro(hoja_trabajo, agrupacion_volumen)
        cls._crear_tabla_dinamica_excel(hoja_trabajo, "TblVolumetria", f"A5:D{hoja_trabajo.max_row}", columna_semaforo="D")

    @classmethod
    def _construir_pestana_calidad_analitica(cls, libro_excel, dataframe):
        """Agrupa las calificaciones, pero esta vez por Métrica (ej. NPS, Eficiencia, Atención)."""
        hoja_trabajo = libro_excel.create_sheet("Control Analítico")
        cls._aplicar_cabecera_institucional(hoja_trabajo, "CONTROL ANALÍTICO DE CALIDAD POR MÉTRICA")
        if dataframe.empty: return
        
        resumen_metricas = dataframe.groupby('Métrica')['Indicador_0_100'].agg(['count', 'mean']).reset_index()
        cls._escribir_dataframe_seguro(hoja_trabajo, resumen_metricas)
        cls._crear_tabla_dinamica_excel(hoja_trabajo, "TblValidacionCalidad", f"A5:C{hoja_trabajo.max_row}", columna_semaforo="C")

    @classmethod
    def _construir_pestana_comentarios_abiertos(cls, libro_excel, dataframe):
        """Filtra y extrae solo los textos largos que dejan los usuarios (Voice of the Customer)."""
        hoja_trabajo = libro_excel.create_sheet("Voz del Cliente")
        cls._aplicar_cabecera_institucional(hoja_trabajo, "ANÁLISIS DE LA VOZ DEL CLIENTE (OPEN TEXT)")
        if dataframe.empty: return
        
        # Máscara: Solo textos con más de 5 caracteres que NO sean respuestas cerradas comunes (SI, NO, N/A)
        filtro_texto_abierto = dataframe['Valor'].apply(lambda x: len(str(x)) > 5 and str(x).upper() not in ['SI', 'NO', 'N/A'])
        dataframe_comentarios = dataframe[filtro_texto_abierto].copy()
        
        cls._escribir_dataframe_seguro(hoja_trabajo, dataframe_comentarios[['Sede', 'Servicio', 'Métrica', 'Valor']])
        
        # Agrandamos bastante la columna D (Valor/Comentario) para que el texto se pueda leer
        hoja_trabajo.column_dimensions['D'].width = 80
        cls._crear_tabla_dinamica_excel(hoja_trabajo, "TblComentariosCliente", f"A5:D{hoja_trabajo.max_row}")

    # ----------------------------------------------------------------------
    # 🛠️ HERRAMIENTAS DE FORMATO INSTITUCIONAL
    # ----------------------------------------------------------------------

    @staticmethod
    def _escribir_dataframe_seguro(hoja_trabajo, dataframe):
        """Escribe un DataFrame forzando todo a texto para evitar que caracteres extraños rompan OpenPyXL."""
        for fila in dataframe_to_rows(dataframe, index=False, header=True):
            # Convertimos a string, pero si el valor es None/Nulo, dejamos la celda vacía ("")
            hoja_trabajo.append([str(celda) if celda is not None else "" for celda in fila])

    @staticmethod
    def _aplicar_cabecera_institucional(hoja_trabajo, titulo_principal):
        """Pinta las primeras dos filas de la hoja con los colores corporativos de la Universidad."""
        # Agrandamos la primera fila para que el logo o texto respire
        hoja_trabajo.row_dimensions[1].height = 45
        hoja_trabajo['A1'] = "UNIVERSIDAD DEL ROSARIO - DITIC"
        hoja_trabajo['A1'].font = Font(bold=True, size=14, color="AF2024") # Rojo
        
        # Combinamos celdas horizontales para que el título ocupe todo el ancho visual
        hoja_trabajo.merge_cells(f'A2:{get_column_letter(max(5, hoja_trabajo.max_column))}2')
        hoja_trabajo['A2'] = titulo_principal.upper()
        hoja_trabajo['A2'].font = Font(bold=True, color="FFFFFF") # Letra blanca
        hoja_trabajo['A2'].fill = PatternFill(start_color="1F4E78", fill_type="solid") # Fondo Azul Oscuro
        hoja_trabajo['A2'].alignment = Alignment(horizontal='center', vertical='center')

    @staticmethod
    def _crear_tabla_dinamica_excel(hoja_trabajo, nombre_tabla_logica, rango_celdas, columna_semaforo=None):
        """Da formato de tabla nativa de Excel y aplica reglas condicionales de semáforo."""
        # Validamos que haya al menos una fila de datos reales (los encabezados están en la 5)
        if hoja_trabajo.max_row < 6: return
        
        # Agregamos una marca de tiempo al nombre de la tabla para evitar choques internos de Excel
        nombre_unico_seguro = f"{nombre_tabla_logica}_{int(time.time()*100)%10000}"
        tabla_excel = Table(displayName=nombre_unico_seguro, ref=rango_celdas)
        tabla_excel.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        hoja_trabajo.add_table(tabla_excel)
        
        # Auto-ajuste de columnas a un mínimo de 20 píxeles
        for columna in hoja_trabajo.columns: 
            hoja_trabajo.column_dimensions[get_column_letter(columna[0].column)].width = 20
        
        # Si se nos pidió semaforizar una columna específica (ej. "E")
        if columna_semaforo:
            # Rango desde la fila 6 (primer dato) hasta el final de la hoja, en la columna dada
            rango_semaforo = f"{columna_semaforo}6:{columna_semaforo}{hoja_trabajo.max_row}"
            # Aplicamos regla de degradado 3 colores
            hoja_trabajo.conditional_formatting.add(
                rango_semaforo, 
                ColorScaleRule(start_type='num', start_value=0, start_color='F8696B', # Rojo
                               mid_type='num', mid_value=75, mid_color='FFEB84', # Amarillo
                               end_type='num', end_value=100, end_color='63BE7B') # Verde
            )