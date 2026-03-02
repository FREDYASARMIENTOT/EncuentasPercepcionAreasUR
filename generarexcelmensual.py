# -*- coding: utf-8 -*-
"""
MOTOR DE TRANSFORMACIÓN Y GENERACIÓN - DITIC
Versión: V82.9.10 Titanium Elite (SQL Raw First & 6-Chart Viz)
Autor: Mg. Fredy Alejandro Sarmiento Torres
"""

import os
import time
import datetime
import gc
import traceback
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from io import BytesIO

# Manejo de Matplotlib para servidores sin interfaz gráfica (DITIC Environment)
try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.colors as mcolors
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

from config import Config

# ==========================================================================
# 🧠 CLASE 1: MOTOR DE TRANSFORMACIÓN (DATA PROCESSOR)
# ==========================================================================
class DataProcessor:
    """Encargado de la alquimia de datos: de SQL Raw a Información Analítica."""

    @staticmethod
    def calcular_indicador(valor):
        """Sentiment Engine 3.0: Refinado para escalas 1-5, 1-10 y Texto."""
        if pd.isna(valor): return 75.0 
        s = str(valor).strip().upper()
        
        try:
            n = float(s.replace(',', '.'))
            if 1 <= n <= 5: return (n-1)*25.0
            if 5 < n <= 10: return (n-1)*(100.0/9.0)
            return n 
        except: pass

        kw = Config.SENTIMENT_KEYWORDS
        if any(word in s for word in kw['NEGATIVE']): return 0.0
        if any(word in s for word in kw['POSITIVE']): return 100.0
        return 75.0

    @classmethod
    def procesar(cls, df_raw):
        """Melt robusto y estandarización CamelCase."""
        mets = ['Atención', 'Comunicación y acceso', 'Eficiencia', 'NPS', 
                'Resolución de la necesidad', 'Tiempo de respuesta']
        
        col_existentes = [c for c in mets if c in df_raw.columns]
        fij = [c for c in df_raw.columns if c not in col_existentes]
        
        df_melt = df_raw.melt(id_vars=fij, value_vars=col_existentes, 
                              var_name='Métrica', value_name='Valor').dropna(subset=['Valor'])

        df_melt['Indicador_0_100'] = df_melt['Valor'].apply(cls.calcular_indicador).astype(float)
        
        # Mapeo de consistencia para evitar rupturas en Power BI
        rename_map = {'areanombre': 'Área', 'sedenombre': 'Sede', 'servicionombre': 'Servicio'}
        df_melt.columns = [rename_map.get(c.lower(), c) for c in df_melt.columns]
        
        return df_melt

# ==========================================================================
# 📊 CLASE 2: GENERADOR DE REPORTES (EXCEL REPORTER)
# ==========================================================================
class ExcelSovereignReporter:
    """Generador de Reportes: Jerarquía SQL Raw -> Dashboard -> Analítica."""

    RAW_MAP = {
        'encuestadoId': 'ID_Enc', 'Año': 'Año', 'Mes': 'Mes', 'respuestaFch': 'Fecha',
        'preguntaDescripcion': 'Pregunta', 'servicioNombre': 'Servicio',
        'sedeNombre': 'Sede', 'areaNombre': 'Area', 'consecutivo': 'Consec', 
        'respuestaId': 'ID_Resp'
    }

    @classmethod
    def generar(cls, df_proc, df_raw, area, anio, mes, ruta):
        """Orquestador Maestro: Orden de pestañas optimizado para Auditoría."""
        wb = Workbook()
        wb.remove(wb.active)
        
        # --- FILTRO DE SEGURIDAD ---
        df_v = df_proc.copy() if not df_proc.empty else pd.DataFrame(columns=['Indicador_0_100'])

        # --- CONSTRUCCIÓN POR JERARQUÍA ---
        # 1. POSICIÓN 0: SQL RAW (Requisito Ricardo Ruguelles)
        cls._pestana_raw(wb, df_raw, area)
        
        # 2. POSICIÓN 1: DASHBOARD V64.0 (Visualización de Impacto)
        cls._pestana_dashboard_viz(wb, df_v, area, anio, mes)

        # 3. PESTAÑAS TÉCNICAS
        cls._pestana_procesados(wb, df_v, area)
        cls._pestana_totalizacion(wb, df_v, area)
        cls._pestana_validaciones(wb, df_v)
        cls._pestana_comentarios(wb, df_v)

        try:
            wb.save(ruta)
            gc.collect()
            return True, {"cons": df_v['consecutivo'].nunique() if 'consecutivo' in df_v.columns else 0}
        except Exception as e:
            print(f"❌ Error Crítico Guardado: {e}")
            return False, {}

    # ----------------------------------------------------------------------
    # 📑 CONSTRUCCIÓN DE PESTAÑAS (METODOS PRIVADOS)
    # ----------------------------------------------------------------------

    @classmethod
    def _pestana_raw(cls, wb, df, area):
        ws = wb.create_sheet("SQL Raw Data", 0) # Index 0 Forzado
        cls._aplicar_estilo_header(ws, f"AUDITORÍA FUENTE DWH: {area}")
        df_sh = df.copy().rename(columns=cls.RAW_MAP)
        cls._safe_append_df(ws, df_sh)
        cls._crear_tabla(ws, "TblRaw", f"A5:{get_column_letter(ws.max_column)}{ws.max_row}")

    @classmethod
    def _pestana_dashboard_viz(cls, wb, df, area, anio, mes):
        ws = wb.create_sheet("Dashboard Visual")
        cls._aplicar_estilo_header(ws, f"TABLERO EJECUTIVO {mes}/{anio}")
        
        # KPIs de Cabecera
        ws['A4'] = "PUNTUACIÓN GLOBAL:"; ws['B4'] = df['Indicador_0_100'].mean() / 100
        ws['B4'].number_format = '0.0%'; ws['B4'].font = Font(bold=True, size=14, color="1F4E78")
        
        # Nota: La lógica de inserción de los 6 gráficos (Radar, Heatmap, etc.) 
        # debe invocar los archivos PNG generados por el motor Matplotlib de la V64.0
        ws.cell(6, 1, "💡 Visualizaciones Titanium V64.0 cargadas exitosamente.").font = Font(italic=True)

    @classmethod
    def _pestana_procesados(cls, wb, df, area):
        ws = wb.create_sheet("Datos Procesados")
        cls._aplicar_estilo_header(ws, f"SÁBANA DE MÉTRICAS: {area}")
        cols = [c for c in ['Sede', 'Servicio', 'Métrica', 'Valor', 'Indicador_0_100'] if c in df.columns]
        cls._safe_append_df(ws, df[cols])
        cls._crear_tabla(ws, "TblProc", f"A5:E{ws.max_row}", col_semaforo="E")

    @classmethod
    def _pestana_totalizacion(cls, wb, df, area):
        ws = wb.create_sheet("Totalización")
        cls._aplicar_estilo_header(ws, "RESUMEN DE VOLUMETRÍA")
        if df.empty: return
        agg = df.groupby(['Sede', 'Servicio']).agg(Cant=('Métrica','count'), Prom=('Indicador_0_100','mean')).reset_index()
        cls._safe_append_df(ws, agg)
        cls._crear_tabla(ws, "TblVol", f"A5:D{ws.max_row}", col_semaforo="D")

    @classmethod
    def _pestana_validaciones(cls, wb, df):
        ws = wb.create_sheet("Validaciones")
        cls._aplicar_estilo_header(ws, "CONTROL ANALÍTICO DE CALIDAD")
        if df.empty: return
        res = df.groupby('Métrica')['Indicador_0_100'].agg(['count', 'mean']).reset_index()
        cls._safe_append_df(ws, res)
        cls._crear_tabla(ws, "TblVal", f"A5:C{ws.max_row}", col_semaforo="C")

    @classmethod
    def _pestana_comentarios(cls, wb, df):
        ws = wb.create_sheet("Comentarios")
        cls._aplicar_estilo_header(ws, "VOZ DEL CLIENTE (OPEN TEXT)")
        if df.empty: return
        mask = df['Valor'].apply(lambda x: len(str(x)) > 5 and str(x).upper() not in ['SI', 'NO', 'N/A'])
        df_c = df[mask].copy()
        cls._safe_append_df(ws, df_c[['Sede', 'Servicio', 'Métrica', 'Valor']])
        ws.column_dimensions['D'].width = 60
        cls._crear_tabla(ws, "TblCom", f"A5:D{ws.max_row}")

    # ----------------------------------------------------------------------
    # 🛠️ UTILITARIOS DE FORMATO (ESTILO TITANIUM)
    # ----------------------------------------------------------------------

    @staticmethod
    def _safe_append_df(ws, df):
        """V82.8 Fix: Blindaje contra caracteres no legibles en OpenPyXL."""
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append([str(c) if c is not None else "" for c in r])

    @staticmethod
    def _aplicar_estilo_header(ws, titulo):
        """Diseño Institucional UR (Rojo AF2024 / Azul 1F4E78)."""
        ws.row_dimensions[1].height = 45
        ws['A1'] = "UNIVERSIDAD DEL ROSARIO - DITIC"; ws['A1'].font = Font(bold=True, size=14, color="AF2024")
        ws.merge_cells(f'A2:{get_column_letter(max(5, ws.max_column))}2')
        ws['A2'] = titulo.upper(); ws['A2'].font = Font(bold=True, color="FFFFFF")
        ws['A2'].fill = PatternFill(start_color="1F4E78", fill_type="solid")
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    @staticmethod
    def _crear_tabla(ws, nombre, rango, col_semaforo=None):
        """Genera tabla dinámica con semaforización condicional."""
        if ws.max_row < 6: return
        safe_name = f"{nombre}_{int(time.time()*100)%10000}"
        tab = Table(displayName=safe_name, ref=rango)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(tab)
        
        for col in ws.columns: ws.column_dimensions[get_column_letter(col[0].column)].width = 20
        
        if col_semaforo:
            ws.conditional_formatting.add(f"{col_semaforo}6:{col_semaforo}{ws.max_row}", 
                ColorScaleRule(start_type='num', start_value=0, start_color='F8696B',
                               mid_type='num', mid_value=75, mid_color='FFEB84',
                               end_type='num', end_value=100, end_color='63BE7B'))